import os
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def categorize_files_to_excel(input_dir: str, output_file: str) -> None:
    """Convert compar TSV files into a single Excel workbook.

    Each TSV becomes its own worksheet, tab‑coloured by basic QC rules.
    """

    # Tab colours
    COLOR_STANDARD = "ffc000"  # yellow‑orange
    COLOR_INVALID = "ff0000"   # red
    COLOR_EMPTY = "92d050"     # green

    # New workbook, ditch the default sheet
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for fname in os.listdir(input_dir):
        if not fname.endswith(".tsv"):
            continue

        fpath = os.path.join(input_dir, fname)
        parts = fname.split(".")
        tab_name = parts[2] if len(parts) > 2 else fname[:31]
        tab_name = tab_name[:31]  # Excel cap

        try:
            df = pd.read_csv(fpath, sep="\t")

            # Decide tab colour
            if df.empty:
                tab_colour = COLOR_EMPTY
            elif "INVALID" in df.to_string(index=False):
                tab_colour = COLOR_INVALID
            else:
                tab_colour = COLOR_STANDARD

            # Write the individual sheet
            ws = wb.create_sheet(title=tab_name)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            ws.sheet_properties.tabColor = tab_colour

        except Exception as exc:
            print(f"Error processing '{fname}': {exc}")

    wb.save(output_file)
    print(f"Excel file '{output_file}' written.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Organise compar TSV files into a coloured Excel workbook.",
    )
    parser.add_argument("input_dir", help="Directory containing compar TSV files")
    parser.add_argument("output_file", help="Output Excel file (.xlsx)")
    args = parser.parse_args()

    categorize_files_to_excel(args.input_dir, args.output_file)
